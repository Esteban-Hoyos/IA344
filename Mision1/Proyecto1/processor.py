import re
from openpyxl import load_workbook
# ===============================
# Función clead_id
#Elimina caracteres no númericos del documento
# ===============================
def clean_id(value):
    if value is None:
        return ""
    return re.sub(r"\D","",str(value))

# ===============================
# Funcion merge_name
#Un nombre y apellido en un solo campo
# ===============================

def merge_name(name,lastname):
    if name is None:
        name=""
    if lastname is None:
        lastname=""
    return f"{name} {lastname}".strip()

def process_excel(path):
    wb=load_workbook(path)
    #Acceso a la hoja llamada "datos"
    ws = wb ["Datos"]
    #Recorrer todas las filas desde la fila 
    for row in range (2,ws.max_row+1):
        #Columna D: identificador limpio
        ws[f"D{row}"] = clean_id (ws[f"A{row}"].value)
        #Columna E: nombre completo
        ws[f"E{row}"] = merge_name(
        ws[f"B{row}"].value,
        ws[f"c{row}"].value
         )
    #Guardar los cambios en el mismo archivo
    wb.save(path)

def process_excel_safe(path):
    try:
        process_excel(path)
        return True, "archivo procesado con éxito"
    except PermissionError:
        return(
            False,
            "El archivo Excel está abierto.\n"
            "Por favor, cierre el archvio e intente nuevamente"
        )
    except KeyError:
        return False, "Hoja `Datos` no encontrada"
    except Exception as e:
        return False,f"Error inesperado:{str(e)}"
    



